"""Excel (.xlsx) measurement sheet: build a fill-in template from the current
config, and parse a filled sheet back into a list of parameter patches.

Design notes:
- The sheet carries a machine-readable "パラメータID" column = the dotted config
  path, so `parse_sheet` matches rows by ID (robust to row reordering/insertion),
  never by row position.
- Only the "計測値(記入)" column is unlocked; the sheet is protected so the user
  can only type into that column.
- Parsing is stateless: it returns {applied, skipped, warnings}. The frontend
  merges `applied` into its live config (reusing its existing deepMerge), so the
  backend never has to hold config state.
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Protection, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from .manifest import MEASURABLE, BY_PATH, SHEET_SCHEMA, Param, get_by_path

HEADERS = ["グループ", "パラメータID(編集不可)", "項目", "単位", "現在値(参考)", "計測値(記入)", "計測方法メモ"]
COL_GROUP, COL_ID, COL_LABEL, COL_UNIT, COL_CURRENT, COL_MEASURED, COL_HOWTO = range(1, 8)
HEADER_ROW = 3
FIRST_DATA_ROW = 4
SHEET_NAME = "計測記入シート"

_TRUE = {"true", "1", "yes", "y", "はい", "on", "有", "有り", "あり", "○", "◯"}
_FALSE = {"false", "0", "no", "n", "いいえ", "off", "無", "無し", "なし", "×", "✕"}


# --------------------------------------------------------------------------- #
# Build
# --------------------------------------------------------------------------- #
def _disp(v: Any, p: Param) -> Any:
    if v is None:
        return None
    if p.kind == "bool":
        return "TRUE" if bool(v) else "FALSE"
    return v


def build_sheet(config: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    widths = {COL_GROUP: 14, COL_ID: 36, COL_LABEL: 24, COL_UNIT: 8,
              COL_CURRENT: 14, COL_MEASURED: 16, COL_HOWTO: 48}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F2937")    # slate-800
    header_font = Font(bold=True, color="FFFFFF")
    banner_fill = PatternFill("solid", fgColor="E5E7EB")    # neutral-200
    banner_font = Font(bold=True, color="111827")
    id_font = Font(color="9CA3AF", italic=True, size=9)     # muted (locked, for parser)
    measured_fill = PatternFill("solid", fgColor="FEF9C3")  # amber-100 "write here" hint
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=COL_HOWTO)
    t = ws.cell(row=1, column=1, value="CSL Simulator 実機計測 記入シート")
    t.font = Font(bold=True, size=14)
    t.alignment = Alignment(horizontal="left", vertical="center")

    # Instructions + machine-readable schema marker
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=COL_HOWTO)
    ins = ws.cell(
        row=2, column=1,
        value=("「計測値(記入)」列だけを編集してください（他セルは保護）。"
               "空欄はスキップされ現在値を維持します。  SHEET_SCHEMA=%d" % SHEET_SCHEMA),
    )
    ins.font = Font(color="6B7280", size=9)
    ins.alignment = Alignment(horizontal="left", vertical="center")

    # Header row (frozen)
    for c, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=c, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    ws.freeze_panes = f"A{FIRST_DATA_ROW}"

    row = FIRST_DATA_ROW
    last_group = None
    for p in MEASURABLE:
        if p.group != last_group:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COL_HOWTO)
            b = ws.cell(row=row, column=1, value=p.group)
            b.fill = banner_fill
            b.font = banner_font
            b.alignment = Alignment(horizontal="left", vertical="center")
            last_group = p.group
            row += 1

        cur = get_by_path(config, p.path)
        ws.cell(row=row, column=COL_GROUP, value=p.group).border = border
        idc = ws.cell(row=row, column=COL_ID, value=p.path); idc.font = id_font; idc.border = border
        lc = ws.cell(row=row, column=COL_LABEL, value=p.label_ja); lc.alignment = left; lc.border = border
        uc = ws.cell(row=row, column=COL_UNIT, value=p.unit); uc.alignment = center; uc.border = border
        cc = ws.cell(row=row, column=COL_CURRENT, value=_disp(cur, p)); cc.alignment = center; cc.border = border
        mc = ws.cell(row=row, column=COL_MEASURED, value=None)
        mc.protection = Protection(locked=False)   # the ONLY editable column
        mc.fill = measured_fill
        mc.alignment = center
        mc.border = border
        hc = ws.cell(row=row, column=COL_HOWTO, value=p.how_to); hc.alignment = left; hc.border = border

        coord = f"{get_column_letter(COL_MEASURED)}{row}"
        dv = None
        if p.kind == "bool":
            dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
        elif p.vmin is not None and p.vmax is not None:
            dv = DataValidation(
                type="whole" if p.kind == "int" else "decimal",
                operator="between", formula1=p.vmin, formula2=p.vmax,
                allow_blank=True, showErrorMessage=True, errorStyle="warning",
                errorTitle="範囲外", error=f"推奨範囲 {p.vmin}〜{p.vmax} 外です（続行できます）。",
            )
        if dv is not None:
            ws.add_data_validation(dv)
            dv.add(coord)

        row += 1

    # Instructions sheet (second tab)
    ws2 = wb.create_sheet("説明")
    ws2.column_dimensions["A"].width = 100
    notes = [
        "CSL Simulator 実機計測 記入シート — 使い方",
        "",
        "1. 「計測記入シート」タブの『計測値(記入)』列（黄色）にだけ実測値を入力します。",
        "2. 分かる項目だけでOKです。空欄の項目は取り込み時にスキップされ、現在値が維持されます。",
        "3. 『はい/いいえ』項目はドロップダウンから TRUE / FALSE を選びます。",
        "4. 数値は推奨範囲を外れると警告が出ますが、そのまま入力・取り込みできます。",
        "5. 『パラメータID』列と行の並びは編集しないでください（取り込みの照合に使います）。",
        "6. 記入後 .xlsx のまま保存し、アプリの『シート取込』ボタンから読み込みます。",
        "",
        "※ 単位に注意（長さ mm / 容積 L / 温度 K / 圧力 Pa / 発熱量 J/kg / 密度 kg/m³）。",
    ]
    for i, line in enumerate(notes, start=1):
        c = ws2.cell(row=i, column=1, value=line)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        if i == 1:
            c.font = Font(bold=True, size=13)

    # Protect the data sheet: only unlocked (measured) cells are editable.
    ws.protection.enable()

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# Parse
# --------------------------------------------------------------------------- #
def _kind_ja(kind: str) -> str:
    return {"bool": "はい/いいえ", "int": "整数", "float": "数値"}.get(kind, "数値")


def _cast(raw: Any, p: Param) -> Any:
    if p.kind == "bool":
        if isinstance(raw, bool):
            return raw
        s = str(raw).strip().lower()
        if s in _TRUE:
            return True
        if s in _FALSE:
            return False
        raise ValueError(s)
    # numeric
    if isinstance(raw, bool):
        raise ValueError("bool in numeric field")
    if isinstance(raw, (int, float)):
        num = float(raw)
    else:
        s = str(raw).strip()
        try:
            num = float(s)
        except ValueError:
            num = float(s.replace(",", "."))   # tolerate decimal-comma locales
    if p.kind == "int":
        return int(round(num))
    return num


def _norm_old(old: Any, p: Param) -> Any:
    if old is None:
        return None
    if p.kind == "bool":
        s = str(old).strip().lower()
        if s in _TRUE:
            return True
        if s in _FALSE:
            return False
        return None
    if isinstance(old, (int, float)):
        return old
    try:
        return float(str(old).strip())
    except ValueError:
        return None


def parse_sheet(data: bytes) -> dict:
    wb = load_workbook(io.BytesIO(data), data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb[wb.sheetnames[0]]

    # Locate header row + the columns we need by matching known labels.
    header_row = col_id = col_measured = col_current = None
    for r in range(1, min(ws.max_row, 12) + 1):
        labels: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                labels[v.strip()] = c
        if "パラメータID(編集不可)" in labels and "計測値(記入)" in labels:
            header_row = r
            col_id = labels["パラメータID(編集不可)"]
            col_measured = labels["計測値(記入)"]
            col_current = labels.get("現在値(参考)")
            break

    applied: list[dict] = []
    skipped: list[dict] = []
    warnings: list[str] = []

    if header_row is None:
        return {"applied": [], "skipped": [],
                "warnings": ["記入シートの見出し行が見つかりません。配布されたテンプレートに記入してください。"]}

    for r in range(header_row + 1, ws.max_row + 1):
        pid = ws.cell(row=r, column=col_id).value
        if not isinstance(pid, str):
            continue
        p = BY_PATH.get(pid.strip())
        if p is None:
            continue
        raw = ws.cell(row=r, column=col_measured).value
        if raw is None or (isinstance(raw, str) and raw.strip() == ""):
            skipped.append({"path": p.path, "label": p.label_ja})
            continue
        old = ws.cell(row=r, column=col_current).value if col_current else None
        try:
            value = _cast(raw, p)
        except ValueError:
            warnings.append(
                f"{p.label_ja}（{p.path}）: 「{raw}」は{_kind_ja(p.kind)}として読めないためスキップしました。")
            skipped.append({"path": p.path, "label": p.label_ja})
            continue
        if p.kind in ("float", "int") and p.vmin is not None and p.vmax is not None:
            if value < p.vmin or value > p.vmax:
                warnings.append(
                    f"{p.label_ja}: {value}{p.unit} は推奨範囲 {p.vmin}〜{p.vmax} 外です（反映しました）。")
        applied.append({"path": p.path, "value": value, "label": p.label_ja,
                        "group": p.group, "old": _norm_old(old, p)})

    return {"applied": applied, "skipped": skipped, "warnings": warnings}
